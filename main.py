




#import all the libarys i need 
from guizero import *
import openpyxl
import re

#to load up the users worksheet
wb = openpyxl.load_workbook('users.xlsx')
ws=wb.active

#to initalise all the screens
app=App(title="Budget Planner")
app2=Window(app,title="Sign up")
app3=Window(app,title="Login")
app4=Window(app,title="error",height=50,width=300)    
app5=Window(app,title="Error",height=50,width=300)
app6=Window(app,title="start screen")
app7=Window(app,title="Income")
app8=Window(app,title="Error",height=50,width=300)
app9=Window(app,title="Error",height=50,width=300)
app10=Window(app,title="Expenses")
app11=Window(app,title="Error",height=50,width=300)
app12=Window(app,title="show")
app13=Window(app,title="Error",height=50,width=300)
app14=Window(app,title="Error",height=50,width=300)
app15=Window(app,title="goals")
app16=Window(app,title="how close to goal")


#a function that gets called when you press a button to take you to the sign up screen
def signUp():
    app4.hide()
    app.hide()
    app2.show()
    app5.hide()
    app8.hide()


#a function that will save the username and password to the user.xlsx file
def saved():
    global c
    c=ws.cell(row=1,column=3).value
    if c==None:
        c=1
    count2=0
    count=0
    if tb1.value=="" or tb2.value=="":
        app2.hide()
        app4.show()
    while count<1:
        for row_pow in range(1,ws.max_row+1):
            b=ws.cell(row=row_pow,column=1).value
            if tb1.value==b:
                app2.hide()
                app5.show()
                count=1
        count=1
    while count2<1:
        for row_pow2 in range(1,ws.max_row+1):
            a=ws.cell(row=row_pow2,column=2).value
            if tb2.value==a:
                app2.hide()
                app8.show()
                count2=1
        count2=1
    if tb1.value!="":
        ws.cell(row=c,column=1,value=tb1.value)
        ws.cell(row=c,column=2,value=tb2.value)
        c=c+1
        ws.cell(row=1,column=3,value=c)
        wb.save('users.xlsx')
        wb2=openpyxl.Workbook(tb1.value+".xlsx")
        wb2.save(tb1.value+".xlsx")

#a function that when called will close all screens save all files and quit the program  
def close():
    global c
    c=ws.cell(row=1,column=3).value
    if c==None:
        c=1
    ws.cell(row=1,column=3,value=c)
    wb.save('users.xlsx')
    exit()

#a function that will take you to the home screen
def home():
    app5.hide()
    app12.hide()
    app8.hide()
    app10.hide()
    app11.hide()
    app6.hide()
    app7.hide()
    app4.hide()
    app9.hide()
    app2.hide()
    app3.hide()
    app.show()
    tb3.clear()
    tb4.clear()

#a function that will take you to the log in screen
def loginScreen():
    app.hide()
    app2.hide()
    app3.show()

    
#a function that will let the user log in if their username and password are correct
def login():
    t8.clear()
    t8.append(text=tb3.value)
    t8.hide()
    c1=0
    c2=False
    c3=0
    c4=False
    count=0
    count2=0
    for row_pow2 in range(1,ws.max_row+1):
            c1=count
            col=ws.cell(row=row_pow2,column=1).value
            if col==tb3.value:
                c2=True
                count=count+1
                break
            elif col!=tb3.value:
                c2=False
    for row_pow3 in range(1,ws.max_row+1):
        col2=ws.cell(row=row_pow3,column=2).value
        c3=count2
        if col2==tb4.value:
            c4=True
            count2=count2+1
            break
    if(c1==c3) and (c2==c4):
        app6.show()
        app3.hide()
        income_Num=0
        expenses_Num=0
        wb2=openpyxl.load_workbook(t8.value+".xlsx")
        ws2=wb2.active
        for row_pow in range(1,ws2.max_row+1):
            if ws2.cell(row=row_pow,column=1).value is not None and ws2.cell(row=row_pow,column=1).value !="new row":
                income_Num=income_Num+int(ws2.cell(row=row_pow,column=1).value)
            if ws2.cell(row=row_pow,column=4).value is not None:
                expenses_Num=expenses_Num+int(ws2.cell(row=row_pow,column=4).value)
        c=income_Num-expenses_Num
        t3.clear()
        t23.append(text=str(c))
        wb2.save(t8.value+".xlsx")
        t26.clear()
        t26.append(text=ws2.cell(row=1,column=7).value)
        
#a function to take the user to the start screen
def loggedIn():
    app6.show()
    app7.hide()
    app12.hide()
    app10.hide()
    app15.hide()
    income_Num=0
    expenses_Num=0
    wb2=openpyxl.load_workbook(t8.value+".xlsx")
    ws2=wb2.active
    for row_pow in range(1,ws2.max_row+1):
            if ws2.cell(row=row_pow,column=1).value is not None and ws2.cell(row=row_pow,column=1).value !="new row":
                income_Num=income_Num+int(ws2.cell(row=row_pow,column=1).value)
            if ws2.cell(row=row_pow,column=4).value is not None:
                expenses_Num=expenses_Num+int(ws2.cell(row=row_pow,column=4).value)
    c=income_Num-expenses_Num
    t23.clear()
    t23.append(text="balance: "+str(c))
    t26.clear()
    t26.append(text="goal: " + str(ws2.cell(row=1,column=7).value))
    wb2.save(t8.value+".xlsx")

#a function that takes the user to the income screen 
def incomescreen():
    app6.hide()
    app7.show()
    app13.hide()
    app9.hide()
    
#a function that takes the users income date and if inputted description and saves it to a personal .xlsx file of the users
def income():
    if (tb5.value==""or tb6.value==""):
        app7.hide()
        app9.show()
    else:
        if tb6.value.count("/")!=2:
            app7.hide()
            app13.show()
        else:
            day,mounth,year=tb6.value.split("/")
            if len(day)!=2 or len(mounth)!=2 or len(year)!=4:
                app7.hide()
                app13.show()
            else:
                if tb7.value=="":
                    tb7.append(text="No Description")
                    wb2=openpyxl.load_workbook(t8.value+".xlsx")
                    ws2=wb2.active
                for row_p in range(1,ws2.max_row+1):
                    a=ws2.cell(row=row_p,column=1).value
                    c=ws2.cell(row=row_p,column=3).value
                    if a is None or a == "new row":
                        ws2.cell(row=row_p, column=1, value=tb5.value)
                        ws2.cell(row=row_p, column=2, value=tb6.value)
                        if c is None:
                            ws2.cell(row=row_p, column=3, value=tb7.value)
                            ws2.cell(row=row_p + 1, column=1, value="new row")
                            wb2.save(t8.value + ".xlsx")
                            break
            wb2.save(t8.value+".xlsx")
        tb5.clear()
        tb6.clear()
        tb7.clear()


#takes the user to the expenses screen
def expensesscreen():
    app10.show()
    app6.hide()
    app11.hide()
    app14.hide()


#a function that takes the users expenses date and if inputted description and saves it to a personal .xlsx file of the users
def expenses():
    date_pattern=r"^\d{2}/\d{2}/\d{4}$"
    if tb10.value==  "" or tb11.value=="":
        app10.hide()
        app11.show()
    else:
        if tb11.value.count("/")!=2:
            app10.hide()
            app14.show()
        else:
            day,mounth,year=tb11.value.split("/")
            if len(day)==2 and len(mounth)==2 and len(year)==4:
                wb3 = openpyxl.load_workbook(t8.value + ".xlsx")
                ws2 = wb3.active
                for row_p in range(1, ws2.max_row + 1):
                    a = ws2.cell(row=row_p, column=4).value
                    if a is None or a =="":
                        ws2.cell(row=row_p, column=4, value=tb10.value)
                        ws2.cell(row=row_p, column=5, value=tb11.value)
                        ws2.cell(row=row_p, column=6, value=tb12.value or "No Description")
                        wb3.save(t8.value + ".xlsx")
                        break
            else:
                app10.hide()
                app14.show()
    tb10.clear()
    tb11.clear()
    tb12.clear()


#a function that will open the show screen and display all the users income and expenses at once
def show():
    lst.clear()
    lst2.clear()
    app6.hide()
    app12.show()
    wb4=openpyxl.load_workbook(t8.value+".xlsx")
    ws4=wb4.active
    for row_pow in range(1,ws4.max_row+1):
        a=ws4.cell(row=row_pow,column=1).value
        b=ws4.cell(row=row_pow,column=2).value
        c2=ws4.cell(row=row_pow,column=3).value
        d=ws4.cell(row=row_pow,column=4).value
        e=ws4.cell(row=row_pow,column=5).value
        f=ws4.cell(row=row_pow,column=6).value
        if a!=None and b!=None and c2!=None:
            lst.append(a+","+b+","+c2)
        if d!=None and e!=None and f!=None:
            lst2.append(d+","+e+","+f)


#a function that will open the goal screen
def goalscreen():
    app6.hide()
    app15.show()


#a function that displays the users goals on the start screen and also saves it to their file 
def goal():
    wb2=openpyxl.load_workbook(t8.value+".xlsx")
    ws2=wb2.active
    ws2.cell(row=1,column=7,value=tb14.value)
    t26.clear()
    t26.append(text="goal" +str(ws2.cell(row=1,column=7).value))
    wb2.save(t8.value+".xlsx")


#a function that tell the user how close they are from their goal(in percentages) by using their balance 
def howclose():
    wb2=openpyxl.load_workbook(t8.value+".xlsx")
    ws2=wb2.active
    app6.hide()
    app16.show()
    income_Num=0
    expenses_Num=0
    for row_pow in range(1,ws2.max_row+1):
            if ws2.cell(row=row_pow,column=1).value is not None and ws2.cell(row=row_pow,column=1).value !="new row":
                income_Num=income_Num+int(ws2.cell(row=row_pow,column=1).value)
            if ws2.cell(row=row_pow,column=4).value is not None:
                expenses_Num=expenses_Num+int(ws2.cell(row=row_pow,column=4).value)
    c=income_Num-expenses_Num
    b=(int(ws2.cell(row=1,column=7).value)/c)*100
    t27.append(text="You are " + str(b) +"% to your goal")

#all the texts boxes name tb1,2,3 ect,all the texts named t1,2,3,ect,all the lists named lst and lst2 and all the pushbutton named b1,2,3,ect
t1=Text(app,text="Do you have an account")
b1=PushButton(app,text="I do not have a account and would like to join",command=signUp)
b2=PushButton(app,text="i have an acount and would like to log in",command=loginScreen)
exits=PushButton(app,text="Exit",command=close)
t2=Text(app2,text="Type the user nane and password you would like below\nUsername")
tb1=TextBox(app2)
t3=Text(app2,text="Password")
tb2=TextBox(app2,hide_text=True)
b3=PushButton(app2,text="sign up",command=saved)
home1=PushButton(app2,text="Home screen",command=home)
log=PushButton(app2,text="login screen",command=loginScreen)
exit2=PushButton(app2,text="Exit",command=close)
t4=Text(app3,text="type in your username and password below to log in\nusername")
tb3=TextBox(app3)
t5=Text(app3,text="password")
tb4=TextBox(app3,hide_text=True)
b3=PushButton(app3,text="login",command=login)
home2=PushButton(app3,text="Home screen",command=home)
exit3=PushButton(app3,text="Exit",command=close)
t6=Text(app4,text="password or user name must be filled in ")
b5=PushButton(app4,text="Ok",command=signUp)
t7=Text(app5,text="Username is in use")
b6=PushButton(app5,text="Ok",command=signUp)
t9=Text(app7)
b4=PushButton(app6,text="add income",command=incomescreen)
t8=Text(app6)
t10=Text(app7,text="you should type in the amount of income \ngotten and when you got it \nand if you want a description")
t11=Text(app7,text="income")
tb5=TextBox(app7)
t12=Text(app7,text="Date(In fotmat DD/MM/YYYY)")
tb6=TextBox(app7)
t13=Text(app7,text="Description")
tb7=TextBox(app7)
add=PushButton(app7,text="Add the income",command=income)
home3=PushButton(app7,text="Home screen",command=loggedIn)
exit5=PushButton(app7,text="Exit",command=close)
t8=Text(app8,text="password already used")
tb8=PushButton(app8,text="Ok",command=signUp)
t9=Text(app9,text="Must input income and date atleats")
tb9=PushButton(app9,text="Ok",command=incomescreen)
b5=PushButton(app6,text="Add expensives",command=expensesscreen)
b6=PushButton(app6,text="to show all incomes and expenses",command=show)
t20=Text(app11,text="Must input expenses and date atleast")
tb13=PushButton(app11,text="Ok",command=expensesscreen )
t14=Text(app10,text="Add your expenses the \ndate you paid it and if\n you want a description")
t15=Text(app10,text="Expenses")
tb10=TextBox(app10)
t16=Text(app10,text="date(in format DD/MM/YYYY)")
tb11=TextBox(app10)
t17=Text(app10,text="description")
tb12=TextBox(app10)
b7=PushButton(app10,text="Expenses",command=expenses)
home4=PushButton(app10,text="Home screen",command=loggedIn)
t18=Text(app12,text="Income,date \nand description")
lst=ListBox(app12,width=120,height="fill")
t19=Text(app12,text="Expenses,\ndate and description")
lst2=ListBox(app12,width=120,height="fill")
home5=PushButton(app12,text="Home screen",command=loggedIn)
exit6=PushButton(app12,text="Exit",command=close)
t20=Text(app13,text="date must be in format DD/MM/YYYY")
b8=PushButton(app13,text="Ok",command=incomescreen)
t21=Text(app14,text="date must be in format DD/MM/YYYY")
b9=PushButton(app14,text="ok",command=expensesscreen)
b10=PushButton(app6,text="mountly goal",command=goalscreen)
t22=Text(app6,text="Your balance and goal is")
t23=Text(app6)
t25=Text(app15,text="set a goal or change you goal")
tb14=TextBox(app15)
b11=PushButton(app15,text="set goal",command=goal)
t26=Text(app6)
b12=PushButton(app15,text="Home screen",command=loggedIn)
b13=PushButton(app15,text="Exit",command=close)
b14=PushButton(app6,text="See how close you are to your goal",command=howclose)
home6=PushButton(app6,text="start screen",command=home)
exit4=PushButton(app6,text="Exit",command=close)
t27=Text(app16)

#to display the main screen and hide all the other screens so they only pop up after being called on
app5.hide()
app13.hide()
app14.hide()
app12.hide()
app8.hide()
app16.hide()
app15.hide()
app10.hide()
app11.hide()
app6.hide()
app7.hide()
app4.hide()
app9.hide()
app2.hide()
app3.hide()
app.display()